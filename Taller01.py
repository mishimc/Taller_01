# -*- coding: utf-8 -*-
"""
@author: PC
"""
#Librerias
#-------------------------------------------------------------------------------------------
import math 
import openpyxl
import pandas as pd
import sqlite3
import os
import seaborn as sns
import calmap
import matplotlib.pyplot as plt

from sqlite3 import Error
from collections import defaultdict
#-------------------------------------------------------------------------------------------
#Leer el archivo 

def leer_excel(ruta_archivo):
    libro = openpyxl.load_workbook(ruta_archivo)
    hoja = libro.active 
    datos = []
    encabezados = []
    for col in hoja.iter_cols(max_row=1):
        encabezados.append(col[0].value)
    for fila in hoja.iter_rows(min_row=2):
        fila_datos = {}
        for idx, celda in enumerate(fila):
            fila_datos[encabezados[idx]] = celda.value
        datos.append(fila_datos)
    
    print(f"Se leyeron {len(datos)} registros del archivo Excel")
    return datos, encabezados


#datos, columnas = leer_excel("C:\\Users\\PC\\Downloads\\Online Retail.xlsx")
#print("Nombres de columnas:", columnas)



#-------------------------------------------------------------------------------------------
#Limpieza de datos

def encontrar_y_eliminar_duplicados(datos, columnas):
    """Identifica y elimina filas duplicadas basadas en columnas clave."""
    registros_vistos = {}
    duplicados = 0
    datos_unicos = []
    
    for registro in datos:
        # Crear una clave única basada en las columnas clave
        clave = tuple(str(registro[col]) if registro[col] is not None else None for col in columnas)
        
        # Si ya hemos visto esta clave, es un duplicado
        if clave in registros_vistos:
            duplicados += 1
        else:
            registros_vistos[clave] = True
            datos_unicos.append(registro)
    
    print(f"\nSe encontraron {duplicados} registros duplicados")
    print(f"Total registros únicos: {len(datos_unicos)}")
    return datos_unicos, duplicados


def completar_datos(datos):
    """
    Completa valores faltantes:
    - Customer ID vacío → '00000'
    - Description vacío → 'Unspecified'
    """
    for registro in datos:
        # 1. Completar Customer ID
        if 'Customer ID' in registro:
            if (registro['Customer ID'] is None or 
                str(registro['Customer ID']).strip() == '' or
                (isinstance(registro['Customer ID'], float) and 
                 math.isnan(registro['Customer ID']))):
                registro['Customer ID'] = '00000'
        
        # 2. Completar Description (nuevo)
        if 'Description' in registro:
            if (registro['Description'] is None or 
                str(registro['Description']).strip() == '' or
                (isinstance(registro['Description'], float) and 
                 math.isnan(registro['Description']))):
                registro['Description'] = 'Unspecified'
    
    return datos


#--------------------------------------------------------------------------------------------
#Conexion con SQLite

def crear_conexion(db_file):
    """Crear una conexión a la base de datos SQLite"""
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        print(f"Conexión a SQLite establecida (versión {sqlite3.version})")
        return conn
    except Error as e:
        print(e)
    return conn

def crear_tabla(conn, encabezados):
    """Crear tabla en la base de datos si no existe"""
    try:
        cursor = conn.cursor()
        
        # Crear sentencia SQL con tipos dinámicos (usaremos TEXT para todos por simplicidad)
        columnas = ", ".join([f'"{col}" TEXT' for col in encabezados])
        sql = f'''CREATE TABLE IF NOT EXISTS retail (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    {columnas}
                );'''
        
        cursor.execute(sql)
        conn.commit()
        print("Tabla creada o ya existente")
    except Error as e:
        print(e)

def insertar_datos(conn, datos, encabezados):
    """Insertar datos"""
    try:
        cursor = conn.cursor()
        
        # Preparar la sentencia SQL
        columnas = ", ".join([f'"{col}"' for col in encabezados])
        placeholders = ", ".join(["?" for _ in encabezados])
        sql = f"INSERT INTO retail ({columnas}) VALUES ({placeholders})"
        
        # Preparar los datos para inserción
        datos_a_insertar = []
        for registro in datos:
            fila = [registro.get(col, None) for col in encabezados]
            datos_a_insertar.append(tuple(fila))
        
        # Insertar en lotes para mejor performance
        cursor.executemany(sql, datos_a_insertar)
        conn.commit()
        print(f"{cursor.rowcount} registros insertados correctamente")
    except Error as e:
        print(e)

# Flujo principal
if __name__ == "__main__":
    ruta_excel = "C:\\Users\\PC\\Downloads\\Online Retail.xlsx"
    ruta_db = "online_retail.db"  # Base de datos en el mismo directorio
    
    # --- NUEVO CÓDIGO AÑADIDO ---
    # Eliminar la base de datos existente si ya existe
    if os.path.exists(ruta_db):
        os.remove(ruta_db)
        print(f"\nBase de datos existente {ruta_db} eliminada")

    datos, columnas = leer_excel(ruta_excel)
    print("Nombres de columnas:", columnas)

    datos_unicos, total_duplicados = encontrar_y_eliminar_duplicados(datos, columnas)
    datos_limpios = completar_datos(datos_unicos)
    
    # 3. Conexión a SQLite y creación de tabla
    conn = crear_conexion(ruta_db)
    
    if conn is not None:
        crear_tabla(conn, columnas)
        insertar_datos(conn, datos_limpios, columnas)
        conn.close()
        print("Proceso completado exitosamente")

#---------------------------------------------------------------------------------------        
#EDA

#Conexion a la base de datos SQLite
def conectar_db(ruta_db):
    """Establecer conexión con la base de datos SQLite"""
    try:
        conn = sqlite3.connect(ruta_db)
        print(f"Conexión exitosa a {ruta_db}")
        return conn
    except sqlite3.Error as e:
        print(f"Error al conectar a la base de datos: {e}")
        return None

#Leer datos y convertirlos a DataFrame
def leer_datos_a_dataframe(conn, nombre_tabla):
    """Leer una tabla SQLite y convertirla a DataFrame de pandas"""
    try:
        # Consulta SQL para obtener todos los datos
        query = f"SELECT * FROM {nombre_tabla}"
        df = pd.read_sql_query(query, conn)
        
        print(f"\nDatos leídos de la tabla '{nombre_tabla}':")
        print(f"- Total de registros: {len(df)}")
        print(f"- Columnas: {list(df.columns)}")
        
        return df
    except Exception as e:
        print(f"Error al leer datos: {e}")
        return pd.DataFrame()  # Devuelve un DataFrame vacío en caso de error

if __name__ == "__main__":
    ruta_db = "online_retail.db"  
    nombre_tabla = "retail"  
    conexion = conectar_db(ruta_db)
    
    if conexion:
        df_retail = leer_datos_a_dataframe(conexion, nombre_tabla)
        
        if not df_retail.empty:
            print("\nVista previa de los datos:")
            print(df_retail.head())

        conexion.close()    
        
df_retail = df_retail.drop(columns=['id'])  
print(df_retail.info())


#Cambios en formato 
df_retail['InvoiceNo'] = df_retail['InvoiceNo'].astype(str)
df_retail['Quantity'] = pd.to_numeric(df_retail['Quantity'], errors='coerce').fillna(0).astype(int)
df_retail['UnitPrice'] = pd.to_numeric(df_retail['UnitPrice'], errors='coerce').astype(float)

# Verificar los cambios
print(df_retail.info())

def analisis_calidad_datos(df):
    print("\n" + "="*50)
    print("ANÁLISIS DE CALIDAD DE DATOS")
    print("="*50)
    
    # Convertir InvoiceDate a datetime si no lo está
    if not pd.api.types.is_datetime64_any_dtype(df['InvoiceDate']):
        df['InvoiceDate'] = pd.to_datetime(df['InvoiceDate'], errors='coerce')
    
    # 1. Análisis por año
    print("\n1. DISTRIBUCIÓN POR AÑO:")
    if 'InvoiceDate' in df:
        df['Year'] = df['InvoiceDate'].dt.year
        conteo_por_anio = df['Year'].value_counts()
        print(conteo_por_anio.to_string())
    
    # 2. Análisis por país
    print("\n2. DISTRIBUCIÓN POR PAÍS:")
    if 'Country' in df:
        conteo_por_pais = df['Country'].value_counts()
        print(f"- Total países únicos: {len(conteo_por_pais)}")
        print("\nTop 10 países por número de compras:")
        print(conteo_por_pais.head(10).to_string())
    
    # 3. Análisis de UnitPrice
    print("\n3. VALORES EN UnitPrice:")
    precios_cero = df[df['UnitPrice'] == 0]
    precios_negativos = df[df['UnitPrice'] < 0]
    print(f"- Registros con precio = 0: {len(precios_cero)}")
    print(f"- Registros con precio < 0: {len(precios_negativos)}")
    
    # 4. Análisis de Quantity
    print("\n4. VALORES EN Quantity:")
    cantidades_negativas = df[df['Quantity'] < 0]
    print(f"- Registros con cantidad < 0: {len(cantidades_negativas)}")
    
    # 5. Análisis de InvoiceNo (cancelaciones)
    print("\n5. FACTURAS CANCELADAS:")
    cancelaciones = df[df['InvoiceNo'].str.startswith('C', na=False)]
    print(f"- Total facturas canceladas: {len(cancelaciones)}")
    
    # 6. Relación entre cancelaciones y cantidades negativas
    print("\n6. RELACIÓN CANCELACIONES/CANTIDADES NEGATIVAS:")
    cancel_neg = df[df['InvoiceNo'].str.startswith('C', na=False) & (df['Quantity'] < 0)]
    print(f"- Cancelaciones con cantidad negativa: {len(cancel_neg)}")
    
    # 7. Análisis de Description (mejorado)
    print("\n7. ANÁLISIS DE DESCRIPCIONES (Description):")

# Primero normalizamos los datos
    df['Description'] = df['Description'].str.strip()

# 1. Descripciones vacías o nulas (mejorado)
    desc_vacias = df['Description'].isna() | (df['Description'] == '')
    empty_count = desc_vacias.sum()

# 2. Detección mejorada de caracteres especiales
# Regex que busca caracteres especiales al inicio, excluyendo espacios
    special_chars_pattern = r'^[^\w\s]'
    desc_con_especial = df['Description'].str.contains(special_chars_pattern, na=False, regex=True)
    special_count = desc_con_especial.sum()

# 3. Detección de múltiples espacios (calidad adicional)
    multi_spaces = df['Description'].str.contains(r'\s{2,}', na=False, regex=True)
    multi_spaces_count = multi_spaces.sum()
    print(f"- Descripciones que comienzan con caracteres especiales (_, ?, #, etc.): {special_count}")
    print(f"- Descripciones con múltiples espacios consecutivos: {multi_spaces_count}")

# Mostramos ejemplos solo si hay casos
    if special_count > 0:
      print("\nEjemplos reales de descripciones con caracteres especiales iniciales:")
      # Filtramos para mostrar solo casos que realmente tienen caracteres especiales al inicio
      real_special_cases = df[desc_con_especial & ~df['Description'].str.startswith(' ')]
      print(real_special_cases['Description'].head(5).to_string(index=False))
    else:
      print("\nNo se encontraron descripciones que comiencen con caracteres especiales.")

# Ejemplos de descripciones con múltiples espacios (si existen)
    if multi_spaces_count > 0:
     print("\nEjemplos de descripciones con múltiples espacios:")
     print(df[multi_spaces]['Description'].head(3).to_string(index=False))
    
     # 8. Análisis de StockCode (nuevo)
     print("\n8. ANÁLISIS DE STOCKCODE:")
     # StockCode que terminan con una letra (regex: [A-Za-z]$)
     stock_con_letra = df['StockCode'].str.contains(r'[A-Za-z]$', na=False, regex=True)
     print(f"- StockCode que terminan con una letra: {stock_con_letra.sum()}")
     print("\nEjemplos de StockCode con letra final:")
     print(df[stock_con_letra]['StockCode'].head(5).to_string(index=False))
    
    return {
        'conteo_por_anio': conteo_por_anio,
        'conteo_por_pais': conteo_por_pais,
        'precios_cero': precios_cero,
        'precios_negativos': precios_negativos,
        'cantidades_negativas': cantidades_negativas,
        'cancelaciones': cancelaciones,
        'cancelaciones_con_cantidad_neg': cancel_neg,
        'desc_vacias': desc_vacias,
        'desc_con_especial': desc_con_especial,
        'stock_con_letra': stock_con_letra
    }

resultados = analisis_calidad_datos(df_retail)

def imprimir_resultados(resultados):
    print("\n" + "="*60)
    print("RESULTADOS COMPLETOS DEL ANÁLISIS".center(60))
    print("="*60)
    
    # 1. Distribución por año
    print("\n[1] DISTRIBUCIÓN TEMPORAL".ljust(60, '-'))
    print("Transacciones por año:")
    print(resultados['conteo_por_anio'].to_string())
    print(f"\nTotal años registrados: {len(resultados['conteo_por_anio'])}")
    
    # 2. Distribución geográfica
    print("\n[2] DISTRIBUCIÓN GEOGRÁFICA".ljust(60, '-'))
    print(f"Países únicos: {len(resultados['conteo_por_pais'])}")
    print("\nTop 10 países con más transacciones:")
    print(resultados['conteo_por_pais'].head(10).to_string())
    
    # 3. Anomalías en precios
    print("\n[3] ANÁLISIS DE PRECIOS".ljust(60, '-'))
    print(f"Registros con precio cero: {len(resultados['precios_cero'])}")
    print(f"Registros con precio negativo: {len(resultados['precios_negativos'])}")
    
    # 4. Anomalías en cantidades
    print("\n[4] ANÁLISIS DE CANTIDADES".ljust(60, '-'))
    print(f"Registros con cantidad negativa: {len(resultados['cantidades_negativas'])}")
    
    # 5. Cancelaciones
    print("\n[5] ANÁLISIS DE CANCELACIONES".ljust(60, '-'))
    print(f"Facturas canceladas (empiezan con 'C'): {len(resultados['cancelaciones'])}")
    print(f"De ellas con cantidad negativa: {len(resultados['cancelaciones_con_cantidad_neg'])}")
    
    print("\n" + "="*60)
    print("MUESTRAS DE REGISTROS PROBLEMÁTICOS".center(60))
    print("="*60)
    
    # Ejemplos de registros problemáticos
    print("\n[Ejemplo 1] Precios negativos:")
    print(resultados['precios_negativos'][['InvoiceNo', 'StockCode', 'Description', 'UnitPrice']].head(3).to_string())
    
    print("\n[Ejemplo 2] Cancelaciones típicas:")
    print(resultados['cancelaciones'][['InvoiceNo', 'CustomerID', 'Quantity', 'UnitPrice']].head(3).to_string())
    
    print("\n[6] ANÁLISIS DE DESCRIPCIONES".ljust(60, '-'))
    print(f"Descripciones vacías/nulas: {resultados['desc_vacias'].sum()}")
    print(f"Descripciones que comienzan con caracteres especiales: {resultados['desc_con_especial'].sum()}")
    
    # Dentro de imprimir_resultados(), después de la sección [6]:
    print("\n[7] ANÁLISIS DE STOCKCODE".ljust(60, '-'))
    print(f"StockCode que terminan con una letra: {resultados['stock_con_letra'].sum()}")

def adicional(df):
    print("---- Informacion adicional ----")
    
    # Asegurarse de que InvoiceDate es datetime
    df['InvoiceDate'] = pd.to_datetime(df['InvoiceDate'], errors='coerce')

    # 1. Fecha mínima y máxima
    fecha_min = df['InvoiceDate'].min()
    fecha_max = df['InvoiceDate'].max()
    print(f"Fecha mínima: {fecha_min}")
    print(f" Fecha máxima: {fecha_max}")

    # 2. País con mayor y menor número de transacciones
    transacciones_por_pais = df.groupby('Country')['InvoiceNo'].nunique().sort_values(ascending=False)
    pais_mayor = transacciones_por_pais.idxmax()
    pais_menor = transacciones_por_pais.idxmin()
    print(f" País con más transacciones: {pais_mayor} ({transacciones_por_pais.max()} transacciones)")
    print(f" País con menos transacciones: {pais_menor} ({transacciones_por_pais.min()} transacciones)")

    # 3. Precio unitario (UnitPrice)
    precios_positivos = df[df['UnitPrice'] > 0]['UnitPrice']
    precios_negativos = df[df['UnitPrice'] < 0]['UnitPrice']

    max_precio = precios_positivos.max()
    min_precio = precios_positivos.min()
    min_precio_negativo = precios_negativos.min() if not precios_negativos.empty else None

    print(f" Precio unitario más caro (positivo): {max_precio}")
    print(f" Precio unitario más barato (positivo): {min_precio}")
    print(f" Precio unitario negativo más bajo: {min_precio_negativo}")

# Llamar a la función de impresión
imprimir_resultados(resultados)
print(adicional(df_retail))

def visualizar_resultados(resultados):
    plt.figure(figsize=(15, 12))
    
    # 1. Distribución temporal (años) - Mantenido
    plt.subplot(2, 2, 1)
    resultados['conteo_por_anio'].plot(kind='bar', color=['#1f77b4', '#ff7f0e'])
    plt.title('Distribución de Transacciones por Año')
    plt.xlabel('Año')
    plt.ylabel('Cantidad de Transacciones')
    for i, v in enumerate(resultados['conteo_por_anio']):
        plt.text(i, v, f"{v:,}", ha='center', va='bottom')
    
    # 2. Top 10 países (log scale) - Mantenido
    plt.subplot(2, 2, 2)
    top_paises = resultados['conteo_por_pais'].head(10)
    ax = top_paises.plot(kind='bar', logy=True, color=sns.color_palette("Spectral", 10))
    plt.title('Top 10 Países por Transacciones (Escala Log)')
    plt.xticks(rotation=45)
    for i, v in enumerate(top_paises):
        ax.text(i, v*1.1, f"{v:,}", ha='center')
    
    # 3. Composición de cancelaciones - Mantenido
    plt.subplot(2, 2, 3)
    cancelaciones = len(resultados['cancelaciones'])
    cancel_neg = len(resultados['cancelaciones_con_cantidad_neg'])
    otros_neg = len(resultados['cantidades_negativas']) - cancel_neg
    sizes = [cancel_neg, otros_neg, cancelaciones - cancel_neg]
    labels = ['Cancel. con cant. neg.', 'Otras cant. neg.', 'Cancel. sin cant. neg.']
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', colors=['#ff9999','#66b3ff','#99ff99'])
    plt.title('Composición de Cancelaciones y Cantidades Negativas')
    
    # 4. Cantidades negativas vs cancelaciones - Mantenido
    plt.subplot(2, 2, 4)
    cantidades = [len(resultados['cantidades_negativas']), len(resultados['cancelaciones_con_cantidad_neg'])]
    labels = ['Cantidades Negativas', 'Asociadas a Cancelaciones']
    plt.bar(labels, cantidades, color=['#ffcc99', '#99ff99'])
    plt.title('Relación Cantidades Negativas y Cancelaciones')
    for i, v in enumerate(cantidades):
        plt.text(i, v, f"{v:,}", ha='center', va='bottom')
    
    plt.tight_layout()
    plt.show()

# Llamar a la función de visualización
visualizar_resultados(resultados)



#---------------------------------------------------------
#Visualizacion 1

# Calcular correlaciones
corr = df_retail[['Quantity', 'UnitPrice']].corr()

# Heatmap avanzado
plt.figure(figsize=(10, 8))
sns.heatmap(corr, annot=True, cmap='coolwarm', center=0, 
            linewidths=.5, annot_kws={"size": 12}, 
            cbar_kws={'label': 'Coeficiente de Correlación'})

plt.title('Correlación entre Variables Clave', pad=20, fontsize=14)
plt.xticks(rotation=45)
plt.yticks(rotation=0)
plt.tight_layout()
plt.savefig('heatmap_correlacion.png', dpi=300, bbox_inches='tight')
plt.show()


#-----------------------------------------------------------
#Visualizacion 2

# Preparar datos
eventos = df_retail.set_index('InvoiceDate')['Quantity'].resample('D').sum()

plt.figure(figsize=(16, 8))
calmap.yearplot(eventos, year=2011, 
               cmap='YlGn', 
               linewidth=0.5,
               daylabels='MTWTFSS',
               dayticks=[0, 2, 4, 6])
plt.title('Actividad de Ventas 2011', pad=20)
plt.savefig('calendario_ventas.png', dpi=300, bbox_inches='tight')

#-------------------------------------------------------------
#Visualizacion 3




# Visualización 3: Análisis de pedidos y cancelaciones por país

# 1. Preparar los datos
def preparar_datos_paises(df):
    # Verifica columnas necesarias
    required_cols = ['Country', 'InvoiceNo', 'Quantity']
    if not all(col in df.columns for col in required_cols):
        raise ValueError(f"Faltan columnas: {required_cols}")

    # Identifica cancelaciones (facturas que empiezan con 'C')
    df['EsCancelacion'] = df['InvoiceNo'].astype(str).str.startswith('C', na=False)
    
    # Agrupa por país
    datos_paises = df.groupby('Country').agg(
        TotalPedidos=('InvoiceNo', 'nunique'),
        TotalCancelaciones=('EsCancelacion', 'sum'),
        TotalItems=('Quantity', lambda x: x[x > 0].sum()),  # Solo items positivos
        TotalItemsCancelados=('Quantity', lambda x: abs(x[x < 0].sum()))  # Items cancelados (aquí faltaba cerrar paréntesis)
    ).reset_index()

    datos_paises['PorcentajeCancelacion'] = datos_paises.apply(lambda x: (x['TotalCancelaciones'] / x['TotalPedidos'] * 100) if x['TotalPedidos'] > 0 else 0, axis=1)
    
    datos_paises['RatioItemsCancelados'] = datos_paises.apply(
        lambda x: (x['TotalItemsCancelados'] / x['TotalItems'] * 100) if x['TotalItems'] > 0 else 0,
        axis=1
    )

    return datos_paises.sort_values('TotalPedidos', ascending=False)

# Ejecuta la función
datos_paises = preparar_datos_paises(df_retail)
print(datos_paises.head())

datos_paises = preparar_datos_paises(df_retail)
plt.figure(figsize=(14, 8))
# Tomar los top 15 países para mejor visualización
top_paises = datos_paises.head(15)
# Configurar el gráfico
ax = plt.subplot()
bar_width = 0.6
bars_total = ax.bar(top_paises['Country'], top_paises['TotalPedidos'], 
                   width=bar_width, color='#2ecc71', label='Pedidos Completos')
bars_cancel = ax.bar(top_paises['Country'], top_paises['TotalCancelaciones'], 
                    width=bar_width, color='#e74c3c', bottom=0, label='Cancelaciones')
# Personalización
plt.title('Facturas de pedidos y cancelaciones por País (Top 15)', fontsize=16, pad=20)
plt.xlabel('País', fontsize=12)
plt.ylabel('Número de Transacciones', fontsize=12)
plt.xticks(rotation=45, ha='right')
plt.grid(axis='y', linestyle='--', alpha=0.7)

# Añadir valores encima de las barras
for bar in bars_total:
    height = bar.get_height()
    ax.text(bar.get_x() + bar.get_width()/2., height,
            f'{int(height)}',
            ha='center', va='bottom', fontsize=9)

# Leyenda mejorada
ax.legend(loc='upper right', framealpha=1)

plt.tight_layout()
plt.savefig('pedidos_cancelaciones_paises.png', dpi=300, bbox_inches='tight')
plt.show()
